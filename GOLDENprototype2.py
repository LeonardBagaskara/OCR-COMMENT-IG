import pytesseract
import  cv2
from PIL import Image
from openpyxl import Workbook, load_workbook


# alamattesseract
pytesseract.pytesseract.tesseract_cmd = r"C:\\Users\\leona\AppData\\Local\\Programs\\Tesseract-OCR\\tesseract.exe"

image_paths = [
        'C:\\Users\\leona\\Downloads\\DESIGN WEB\\ig\\77.jpg',
        'C:\\Users\\leona\\Downloads\\DESIGN WEB\\ig\\88.jpg',
        'C:\\Users\\leona\\Downloads\\DESIGN WEB\\ig\\99.jpg'
        # untuk menambahkan alamat file gambar yang ingin diproses
    ]
excel_path = 'C:\\Users\\leona\\Downloads\\DESIGN WEB\\ig\\data.xlsx' #input database
output_excel_path = 'C:\\Users\\leona\\Downloads\\DESIGN WEB\\ig\\64tesAKURASI.xlsx'
output_paths = [
        'C:\\Users\\leona\\Downloads\\DESIGN WEB\\ig\\77.jpg',
        'C:\\Users\\leona\\Downloads\\DESIGN WEB\\ig\\88.jpg',
        'C:\\Users\\leona\\Downloads\\DESIGN WEB\\ig\\99.jpg'
    # OUTPUT EDIT IMG
]

def downgrade_image_pixel(image_paths, output_paths, new_size):
    """
    Resize karena ocr max 32100px.

    """
    # Ensure the lists have the same length
    if len(image_paths) != len(output_paths):
        raise ValueError("The lists of image paths and output paths must have the same length.")
    
    # Iterate over each image path and resize the image
    for img_path, out_path in zip(image_paths, output_paths):
        with Image.open(img_path) as img:
            # Resize the image
            resized_img = img.resize(new_size, Image.LANCZOS)
            
            # Save the resized image
            resized_img.save(out_path)

new_size = (1080, 32000) # Adjust the size as needed

downgrade_image_pixel(image_paths, output_paths, new_size)



#NEXT STEP : CROP

def crop_image(image_paths, output_paths, left, upper, right, lower):
    """
    Crop Gambar supaya OCR bekerja secara optimal.

    Parameters:
    - image_path: Alamat dimana file gambar yang akan dicrop di simpan.
    - output_path: Alamat dimana file gambar yang sudah dicrop di simpan.
    - left: x-coordinate dari sisi kiri crop box. 
    - upper: y-coordinate dari sisi atas crop box.
    - right: The x-coordinate of the right side of the crop box.
    - lower: The y-coordinate of the lower side of the crop box.

    """
    # memastikan sama lebar
    if len(image_paths) != len(output_paths):
        raise ValueError("The lists of image paths and output paths must have the same length.")
    
    # crop
    for img_path, out_path in zip(image_paths, output_paths):
        with Image.open(img_path) as img:
            # Crop the image
            cropped_img = img.crop((left, upper, right, lower))
            
            # Save the cropped image
            cropped_img.save(out_path)

left = 150
upper = 325
right = 939
lower = 31700 # Adjust this value as needed based on the resolution of the device used for the screenshot

crop_image(image_paths, output_paths, left, upper, right, lower)


#NEXT STEP PROSESS OCR, FILTER, SAVE TO EXCELL

def read_words_from_excel(excel_path):
    """
    membaca database
    """
    wb = load_workbook(excel_path)
    ws = wb.active
    words = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            cell_str = str(cell) if cell is not None else ""
            words.extend(cell_str.split())
    return words

def ocr_image(image_path, words):
    """
    melakukan ocr dan memfilter hasil
    """
    image = Image.open(image_path)
    text = pytesseract.image_to_string(image)
    
    text_with_spaces = ' '.join(text.split())

    ignore_phrases = ["Reply", "See", "translation", "@", "/"]

    for phrase in ignore_phrases:
        text_with_spaces = text_with_spaces.replace(phrase, '')
        
    filtered_text = ' '.join(word for word in text_with_spaces.lower().split() if word.lower() in words)
    
    words_without_duplicates = list(dict.fromkeys(filtered_text.split()))
    
    text_without_duplicates = ' '.join(words_without_duplicates)
    
    return text_without_duplicates

def write_to_excel(text, excel_path):
    """
    menulis data yang dikenali dari database
    """
    wb = Workbook()
    ws = wb.active
    words = text.split()
    for word in words:
        ws.append([word])
    wb.save(excel_path)

def main():


    # cek database
    words = read_words_from_excel(excel_path)

    # menambahkan spasi diantara kata kata dan mengumpulkan hasil
    all_text = ""

    for image_path in image_paths:
        # ocr dan filter text yang cocok kriteria
        text = ocr_image(image_path, words)
        
        all_text += text + " "

    # save hasil ke file exsel
    write_to_excel(all_text, output_excel_path)

    print(f"Text recognized from all images and saved to '{output_excel_path}'.")

if __name__ == "__main__":
    main()
